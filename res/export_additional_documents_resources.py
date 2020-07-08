from db_models.models import ConsolidateStaticDocuments, ConsolidateExcludeTransactionsDocuments
from db.db import session
from flask_restful import Resource, fields, marshal_with, abort, reqparse
from flask import Flask, make_response, send_from_directory, send_file, request
from modules.original_file_saver import save_original_documents_documents
from pathlib import Path
import pandas as pd
import numpy as np
import json
from modules.documents_exporter import translit, clean_filename
from settings import EXPORT_FOLDER
import os
import uuid
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Alignment,Font,Border,Side
from openpyxl.utils.dataframe import dataframe_to_rows


class ExportExcludeTransactionsDocumentsResource(Resource):
    def get(self, id):
        try:
            document = session.query(ConsolidateExcludeTransactionsDocuments).filter(
                ConsolidateExcludeTransactionsDocuments.id == id).first()

            if (not document):
                return make_response('Files not found', 400)
            data = json.loads(document.data)
            balance_header = data['balance_header']
            balance_data = data['balance_data']
            if (len(balance_data) > 0):
                balance_frame = pd.DataFrame(np.array(balance_data),
                                             columns=balance_header)
            else:
                balance_frame = pd.DataFrame(
                    columns=balance_header)
            balance_frame.drop(columns=['', 'id'], inplace=True)

            opiu_header = data['opiu_header']
            opiu_data = data['opiu_data']
            if (len(opiu_data) > 0):
                opiu_frame = pd.DataFrame(np.array(opiu_data),
                                          columns=opiu_header)
            else:
                opiu_frame = pd.DataFrame(columns=opiu_header)

            opiu_frame.drop(columns=['', 'id'], inplace=True)

            odds_header = data['odds_header']
            odds_data = data['odds_data']
            if (len(odds_data) > 0):
                odds_frame = pd.DataFrame(np.array(odds_data),
                                          columns=odds_header)
            else:
                odds_frame = pd.DataFrame(
                    columns=odds_header)

            odds_frame.drop(columns=['', 'id'], inplace=True)

            dir_id = str(uuid.uuid4().hex)
            project_folder = os.path.join(EXPORT_FOLDER, dir_id)
            if not os.path.exists(project_folder):
                os.makedirs(project_folder)
            name = 'export_exclude_data_' + document.name + '.xlsx'

            name = str(name).replace('"', ' ')
            name = translit(name, 'ru', reversed=True)

            file_name = clean_filename(name)
            file_name = file_name.replace('__', "_")

            export_path = os.path.join(project_folder, file_name)
            wb = Workbook()

            balance_sheet = wb.create_sheet("Баланс")
            index = 1
            for r in dataframe_to_rows(balance_frame, index=True, header=True):
                l = len(r)
                balance_sheet.append(r)
                if (index > 2):
                    for cell_index in range(6,10):
                        value = balance_sheet.cell(index, cell_index).value
                        value = str(value).replace(',', '')
                        # value = str(value).replace('.',',')
                        value = float(value)
                        balance_sheet.cell(index, cell_index).value = value

                        balance_sheet.cell(index, cell_index).number_format = '#,##0'
                        balance_sheet.cell(index, cell_index).alignment = Alignment(horizontal='center')

                index += 1

            balance_sheet.delete_rows(2, 1)
            balance_sheet.delete_cols(1, 1)

            dims = {}
            for row in balance_sheet.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))


            for col, value in dims.items():
                balance_sheet.column_dimensions[col].width = value
            border = Border(left=Side(border_style='thin', color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin', color='000000'),
                            bottom=Side(border_style='thin', color='000000'))
            for rows in balance_sheet.iter_rows(min_row=1, max_row=1, min_col=1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='d4f2ff', end_color='d4f2ff',  fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold = True)
                    cell.border = border


            opiu_sheet = wb.create_sheet("ОПиУ")
            index = 1
            for r in dataframe_to_rows(opiu_frame, index=True, header=True):
                opiu_sheet.append(r)
                if (index>2):
                    for cell_index in range(8, 13):
                        if (cell_index==10 or cell_index==12):
                            value=opiu_sheet.cell(index,cell_index).value
                            value = str(value).replace(',','')

                            value = float(value)
                            opiu_sheet.cell(index, cell_index).value = value
                            opiu_sheet.cell(index, cell_index).alignment = Alignment(horizontal='center')
                            opiu_sheet.cell(index,cell_index).number_format = '#,##0'

                index+=1

            opiu_sheet.delete_rows(2, 1)
            opiu_sheet.delete_cols(1, 1)

            dims = {}


            for row in opiu_sheet.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                        cell.alignment= Alignment(horizontal='center')
            for col, value in dims.items():
                opiu_sheet.column_dimensions[col].width = value

            for rows in opiu_sheet.iter_rows(min_row=1, max_row=1, min_col=1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='d4f2ff', end_color='d4f2ff',  fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold = True)
                    cell.border = border


            odds_sheet = wb.create_sheet("ОДДС")
            index = 1
            for r in dataframe_to_rows(odds_frame, index=True, header=True):
                odds_sheet.append(r)
                if (index>2):
                    for cell_index in range(8, 13):
                        if (cell_index==10 or cell_index==12):
                            value=odds_sheet.cell(index,cell_index).value
                            value = str(value).replace(',','')

                            value = float(value)
                            odds_sheet.cell(index, cell_index).value = value

                            odds_sheet.cell(index,cell_index).number_format = '#,##0'

                index+=1

            odds_sheet.delete_rows(2, 1)
            odds_sheet.delete_cols(1, 1)

            dims = {}
            for row in odds_sheet.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
                        cell.alignment = Alignment(horizontal='center')
            for col, value in dims.items():
                odds_sheet.column_dimensions[col].width = value


            for rows in odds_sheet.iter_rows(min_row=1, max_row=1, min_col=1):
                for cell in rows:
                    cell.fill = PatternFill(start_color='d4f2ff', end_color='d4f2ff',  fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
                    cell.font = Font(bold = True)
                    cell.border = border

            r_sheet = wb.get_sheet_by_name('Sheet')
            wb.remove(r_sheet)
            wb.save(export_path)



            return send_from_directory(project_folder, file_name,
                                       as_attachment=True)

        except Exception as e:
            abort(404, message="File not found")


class ExportStaticDocumentsResource(Resource):
    def post(self):
        try:

            json_data = request.get_json(force=True)
            project_id = json_data["project_id"]
            document_id = json_data["document_id"]

            static_document = session.query(ConsolidateStaticDocuments).filter(
                ConsolidateStaticDocuments.project_id == project_id).order_by(
                ConsolidateStaticDocuments.id.desc()).first()

            balance_static_document_id = static_document.data['balance_static_document']['id']
            opiu_static_document_id = static_document.data['opiu_static_document']['id']
            deb_credit_certificate_id = static_document.data['deb_credit_certificate']['id']

            file_path = ''

            if (balance_static_document_id == document_id):
                file_path = static_document.data['balance_static_document']['file_path']
            elif (opiu_static_document_id == document_id):
                file_path = static_document.data['opiu_static_document']['file_path']
            elif (deb_credit_certificate_id == document_id):
                file_path = static_document.data['deb_credit_certificate']['file_path']

            if (file_path == ''):
                return make_response('Files not found', 400)

            p = Path(file_path)

            folder_path = p.parent
            file_path = p.stem
            file_ext = p.suffix
            name = str(file_path + file_ext).replace(' ', '_')

            name = str(name).replace('"', ' ')
            name = translit(name, 'ru', reversed=True)

            file_name = clean_filename(name)
            file_name = file_name.replace('__', "_")

            return send_from_directory(folder_path, file_path + file_ext, attachment_filename=file_name,
                                       as_attachment=True)
        except Exception as e:
            abort(404, message="File not found")
